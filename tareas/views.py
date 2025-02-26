from rest_framework import generics
from .models import Tareas, AsignacionesTareas,VistaEmpleadosTareas,ObservacionTarea
from .serializers import ObservacionTareaSerializer,TareasSerializer, AsignacionesTareasReadSerializer,AsignacionesTareasCreateSerializer,AsignacionesTareasUpdateSerializer
from rest_framework.views import APIView
from .services import calcular_metricas_por_empleado
from rest_framework.response import Response
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border,Side
from django.db.models import F
from django.utils import timezone
from datetime import timedelta
from rest_framework import viewsets

# Vistas para Tareas
class TareasListCreateView(generics.ListCreateAPIView):
    queryset = Tareas.objects.all()
    serializer_class = TareasSerializer

class TareasRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Tareas.objects.all()
    serializer_class = TareasSerializer

# Vistas para AsignacionesTareas
class AsignacionesTareasListCreateView(generics.ListCreateAPIView):
    queryset = AsignacionesTareas.objects.all()

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return AsignacionesTareasCreateSerializer
        return AsignacionesTareasReadSerializer

class AsignacionesTareasRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = AsignacionesTareas.objects.all()

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return AsignacionesTareasUpdateSerializer
        return AsignacionesTareasReadSerializer
    
class TareasPorEmpleadoView(generics.ListCreateAPIView):
    serializer_class = AsignacionesTareasReadSerializer

    def get_queryset(self):
        empleado_id = self.kwargs['empleado_id']
        return AsignacionesTareas.objects.filter(empleado_id=empleado_id)

class MetricasPorEmpleadoView(APIView):
    def get(self, request, empleado_id, *args, **kwargs):
        metricas = calcular_metricas_por_empleado(empleado_id)
        return Response(metricas)
    
class ReporteTareasNoEntregadasUltimoMesView(APIView):
    def get(self, request, *args, **kwargs):
        # Obtener la fecha actual
        fecha_actual = timezone.now()

        # Filtrar las tareas que ya están vencidas
        tareas_vencidas = VistaEmpleadosTareas.objects.filter(
            fecha_estimada_fin__lt=fecha_actual
        )

        # Crear un libro de trabajo y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tareas Vencidas"

        # Definir los encabezados
        headers = ["ID Empleado", "Nombre Empleado", "Apellido Empleado", "ID Tarea", "Título Tarea", "Descripción Tarea", "Fecha Inicio", "Fecha Estimada Fin", "Fecha Real Fin", "Estado Tarea"]

        # Agregar los encabezados
        ws.append(headers)
        ws.row_dimensions[1].height = 35  # Ajustar la altura de la fila 1
        for cell in ws[1]:
            cell.font = Font(name='Arial Black', size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar y alinear al centro
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

        # Función para eliminar la información de zona horaria
        def remove_tzinfo(dt):
            if dt is not None and dt.tzinfo is not None:
                return dt.replace(tzinfo=None)
            return dt

        # Agregar las tareas a la hoja
        for tarea in tareas_vencidas:
            ws.append([
                tarea.id_empleado, tarea.nombre_empleado, tarea.apellido_empleado, tarea.tarea_id,
                tarea.titulo_tarea, tarea.descripcion_tarea, remove_tzinfo(tarea.fecha_inicio),
                remove_tzinfo(tarea.fecha_estimada_fin), remove_tzinfo(tarea.fecha_real_fin), tarea.estado_tarea
            ])

        # Ajustar el ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_tareas_vencidas.xlsx'
        wb.save(response)
        return response

class ReporteTareasPendientesEnProgresoView(APIView):
    def get(self, request, *args, **kwargs):
        # Filtrar las tareas que están en estado "Pendiente" o "En Progreso"
        tareas_pendientes_en_progreso = VistaEmpleadosTareas.objects.filter(
            estado_tarea__in=["Pendiente", "En Progreso"]
        )

        # Crear un libro de trabajo y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tareas Pendientes o En Progreso"

        # Definir los encabezados
        headers = ["ID Empleado", "Nombre Empleado", "Apellido Empleado", "ID Tarea", "Título Tarea", "Descripción Tarea", "Fecha Inicio", "Fecha Estimada Fin", "Fecha Real Fin", "Estado Tarea"]

        # Agregar los encabezados
        ws.append(headers)
        ws.row_dimensions[1].height = 35  # Ajustar la altura de la fila 1
        for cell in ws[1]:
            cell.font = Font(name='Arial Black', size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar y alinear al centro
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

        # Función para eliminar la información de zona horaria
        def remove_tzinfo(dt):
            if dt is not None and dt.tzinfo is not None:
                return dt.replace(tzinfo=None)
            return dt

        # Agregar las tareas a la hoja
        for tarea in tareas_pendientes_en_progreso:
            ws.append([
                tarea.id_empleado, tarea.nombre_empleado, tarea.apellido_empleado, tarea.tarea_id,
                tarea.titulo_tarea, tarea.descripcion_tarea, remove_tzinfo(tarea.fecha_inicio),
                remove_tzinfo(tarea.fecha_estimada_fin), remove_tzinfo(tarea.fecha_real_fin), tarea.estado_tarea
            ])

        # Ajustar el ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_tareas_pendientes_en_progreso.xlsx'
        wb.save(response)
        return response

class ObservacionTareaViewSet(viewsets.ModelViewSet):
    queryset = ObservacionTarea.objects.all()
    serializer_class = ObservacionTareaSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        tarea_id = self.request.query_params.get('tarea_id', None)
        if tarea_id is not None:
            queryset = queryset.filter(tarea_id=tarea_id)
        return queryset